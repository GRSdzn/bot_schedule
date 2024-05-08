from datetime import timedelta, datetime
from openpyxl import load_workbook
import os
import requests
from dotenv import load_dotenv

load_dotenv()
URL = os.getenv("URL")


def download_schedule():
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (XHTML, like Gecko) '
                      'Chrome/58.0.3029.110 Safari/537.3'}
    r = requests.get(URL, headers=headers, verify=False)
    with open('./schedule.xlsx', 'wb') as f:
        f.write(r.content)


def get_or_download_schedule_file():
    if not os.path.isfile('schedule.xlsx'):
        download_schedule()
    return load_workbook('schedule.xlsx').active


def get_schedule_for_date(date):
    """
    Get the schedule for a specific date.
    Args:
        date: The date for which the schedule is required.
    Returns:
        The schedule for the specified date.
    """
    sheet = get_or_download_schedule_file()
    schedule = []
    today = datetime.now().date()

    for row in sheet.iter_rows(min_row=16, values_only=True):
        raw_date = row[0]
        if raw_date is None:
            continue

        date_parts = raw_date.split('\n')
        raw_date = date_parts[0]

        try:
            date_obj = datetime.strptime(raw_date, '%d.%m').date().replace(year=today.year)
        except ValueError:
            try:
                date_obj = datetime.strptime(raw_date, '%m.%d').date().replace(year=today.year)
            except ValueError:
                continue

        if date_obj == date:
            weekday = date_parts[1]

            subject = row[1]
            teacher = row[2]
            class_type = row[3]
            time = row[4]
            room = row[5]

            schedule.append(
                f"<b><u>Пары на {date_obj:%d.%m} ({weekday[:2]})</u></b>\n"
                f"📖Дисциплина: <b>{subject}</b>\n"
                f"👨‍🏫Преподаватель: <b>{teacher}</b>\n"
                f"🔖Вид: <b>{class_type}</b>\n"
                f"⌚️Время: <b>{time}</b>\n"
                f"🔑Ауд: <b>{room}</b>\n"
            )

    if len(schedule) > 0:
        return '\n'.join(schedule)
    else:
        return f'- <i>Пары отсутствуют в расписании на {date:%d.%m}</i>\n'


def get_today_schedule():
    """
    Get today's schedule.
    Returns:
        Today's schedule.
    """
    today = datetime.now().date()
    return get_schedule_for_date(today)


def get_tomorrow_schedule():
    """
    Get tomorrow's schedule.
    Returns:
        Tomorrow's schedule.
    """
    today = datetime.now().date()
    tomorrow = today + timedelta(days=1)
    return get_schedule_for_date(tomorrow)


def get_week_schedule():
    """
    Get the schedule for the week.
    Returns:
        The schedule for the entire week.
    """
    today = datetime.now().date()
    start_date = today + timedelta(days=(7 - today.weekday()) % 7)

    schedule = []
    current_date = today

    for _ in range(7):
        schedule.append(get_schedule_for_date(current_date))
        current_date += timedelta(days=1)

    return '\n'.join(schedule)
